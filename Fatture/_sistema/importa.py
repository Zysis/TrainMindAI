#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Sistema spese TrainMind — importatore fatture.

Cosa fa:
  1. Scansiona le sottocartelle di Fatture/ cercando PDF nuovi
  2. Estrae fornitore, numero, data, imponibile, IVA, totale, valuta, periodo
  3. Normalizza tutto in EUR e salva in dati/spese.json (fonte di verita')
  4. Unisce le spese inserite a mano (spese_manuali.csv)
  5. Rigenera Registro-Spese.xlsx e Dashboard-Spese.html

Uso:
    python _sistema/importa.py                  # fa tutto: metriche VPS, PDF, Excel, dashboard
    python _sistema/importa.py --senza-metriche # salta il collegamento al VPS
    python _sistema/importa.py --solo-report    # rigenera solo Excel e dashboard
    python _sistema/importa.py --reimporta      # rilegge tutti i PDF da zero

Dipendenze: python3, openpyxl, pdftotext (poppler-utils). Per le metriche: ssh.
"""

from __future__ import annotations

import argparse
import csv
import hashlib
import json
import os
import re
import shutil
import subprocess
import sys
from collections import defaultdict
from datetime import date, datetime, timedelta
from pathlib import Path

BASE = Path(__file__).resolve().parent.parent      # cartella Fatture/
SISTEMA = BASE / "_sistema"
DATI = BASE / "dati"
CONFIG_PATH = SISTEMA / "config.json"
SPESE_JSON = DATI / "spese.json"
METRICHE_JSON = DATI / "metriche.json"
MANUALI_CSV = BASE / "spese_manuali.csv"
SQL_METRICHE = SISTEMA / "metriche.sql"
XLSX_OUT = BASE / "Registro-Spese.xlsx"
HTML_OUT = BASE / "Dashboard-Spese.html"

CARTELLE_IGNORATE = {"_sistema", "dati", "__pycache__", ".git"}

MESI_EN = {
    "january": 1, "february": 2, "march": 3, "april": 4, "may": 5, "june": 6,
    "july": 7, "august": 8, "september": 9, "october": 10, "november": 11, "december": 12,
}
MESI_IT = ["", "Gennaio", "Febbraio", "Marzo", "Aprile", "Maggio", "Giugno",
           "Luglio", "Agosto", "Settembre", "Ottobre", "Novembre", "Dicembre"]


# --------------------------------------------------------------------------
# utilita'
# --------------------------------------------------------------------------

def num_it(s: str | None) -> float | None:
    """'1.234,56' -> 1234.56"""
    if s is None:
        return None
    s = s.strip().replace(" ", "").replace(" ", "")
    if not s:
        return None
    s = s.replace(".", "").replace(",", ".")
    try:
        return float(s)
    except ValueError:
        return None


def num_en(s: str | None) -> float | None:
    """'1,234.56' -> 1234.56"""
    if s is None:
        return None
    s = s.strip().replace(" ", "").replace(" ", "").replace("$", "").replace("€", "")
    if not s:
        return None
    s = s.replace(",", "")
    try:
        return float(s)
    except ValueError:
        return None


def iso_da_it(s: str | None) -> str | None:
    """'15/07/2026' -> '2026-07-15'"""
    if not s:
        return None
    try:
        return datetime.strptime(s.strip(), "%d/%m/%Y").date().isoformat()
    except ValueError:
        return None


def iso_da_en(s: str | None) -> str | None:
    """'July 21, 2026' -> '2026-07-21'"""
    if not s:
        return None
    m = re.match(r"([A-Za-z]+)\s+(\d{1,2}),\s*(\d{4})", s.strip())
    if not m:
        return None
    mese = MESI_EN.get(m.group(1).lower())
    if not mese:
        return None
    return date(int(m.group(3)), mese, int(m.group(2))).isoformat()


def sha_file(p: Path) -> str:
    h = hashlib.sha256()
    with p.open("rb") as f:
        for blocco in iter(lambda: f.read(65536), b""):
            h.update(blocco)
    return h.hexdigest()[:16]


def pdf_testo(p: Path) -> str:
    if not shutil.which("pdftotext"):
        raise RuntimeError(
            "pdftotext non trovato. Installa poppler-utils "
            "(Linux: apt install poppler-utils · Windows: scoop install poppler)."
        )
    res = subprocess.run(
        ["pdftotext", "-layout", str(p), "-"],
        capture_output=True, text=True, encoding="utf-8", errors="replace",
    )
    if res.returncode != 0:
        raise RuntimeError(f"pdftotext ha fallito su {p.name}: {res.stderr[:200]}")
    return res.stdout


# --------------------------------------------------------------------------
# parser per fornitore
# --------------------------------------------------------------------------

def parse_ionos(t: str) -> dict:
    d: dict = {"valuta": "EUR", "avvisi": []}

    m = re.search(r"Numero fattura:\s+(\d+)", t)
    d["numero"] = m.group(1) if m else None

    m = re.search(r"Data fattura:\s+(\d{2}/\d{2}/\d{4})", t)
    d["data"] = iso_da_it(m.group(1)) if m else None

    m = re.search(r"Contratto:\s*(\d+)\s*-\s*(.+)", t)
    if m:
        d["contratto"] = m.group(1).strip()
        d["descrizione"] = m.group(2).strip()

    # i totali compaiono sull'ultima pagina: prendo sempre l'ultima occorrenza
    sub = re.findall(r"Subtotale \(netto\)\s+([\d.,]+)\s*EUR", t)
    d["netto"] = num_it(sub[-1]) if sub else None

    iva = re.findall(r"\+\s*IVA\s*\(([\d,]+)\s*%\)\s+([\d.,]+)\s*EUR", t)
    if iva:
        d["iva_percento"] = num_it(iva[-1][0])
        d["iva"] = num_it(iva[-1][1])

    tot = re.findall(r"Totale da pagare\s+([\d.,]+)\s*EUR", t)
    d["totale"] = num_it(tot[-1]) if tot else None

    # periodo coperto dalla fattura (prima riga "gg/mm/aaaa-gg/mm/aaaa")
    m = re.search(r"(\d{2}/\d{2}/\d{4})-(\d{2}/\d{2}/\d{4})", t)
    if m:
        d["periodo_da"] = iso_da_it(m.group(1))
        d["periodo_a"] = iso_da_it(m.group(2))

    # scadenza di eventuali sconti promozionali
    m = re.search(r"Valido da (\d{2}/\d{2}/\d{4}) a (\d{2}/\d{2}/\d{4})", t)
    if m:
        d["sconto_da"] = iso_da_it(m.group(1))
        d["sconto_a"] = iso_da_it(m.group(2))

    # voci di dettaglio
    voci = []
    for riga in t.splitlines():
        m = re.match(r"^\s*(\d{1,2})\s{2,}(\S.*?)\s{2,}.*?(-?[\d.,]+)\s+([\d,]+)\s*%\s*$", riga)
        if m:
            voci.append({
                "n": int(m.group(1)),
                "descrizione": re.sub(r"\s{2,}", " ", m.group(2)).strip(),
                "netto": num_it(m.group(3)),
                "iva_percento": num_it(m.group(4)),
            })
    d["voci"] = voci

    return d


def parse_openai(t: str) -> dict:
    d: dict = {"valuta": "USD", "avvisi": []}

    m = re.search(r"Invoice number\s+(\S+)", t)
    d["numero"] = m.group(1) if m else None

    m = re.search(r"Date of issue\s+([A-Za-z]+\s+\d{1,2},\s*\d{4})", t)
    d["data"] = iso_da_en(m.group(1)) if m else None

    m = re.search(r"Subtotal\s+\$?([\d,.]+)", t)
    d["netto"] = num_en(m.group(1)) if m else None

    # l'importo IVA sta su una riga isolata poco prima della riga "VAT ..."
    iva_usd = None
    righe = t.splitlines()
    for i, riga in enumerate(righe):
        if re.search(r"\bVAT\b.*\(\s*[\d.,]+\s*%", riga):
            m2 = re.search(r"IVA[^\d]*([\d.,]+)\s*%|(\d+(?:[.,]\d+)?)\s*%", riga)
            for j in range(max(0, i - 3), min(len(righe), i + 4)):
                m3 = re.search(r"^\s*\$([\d,.]+)\s*$", righe[j])
                if m3:
                    iva_usd = num_en(m3.group(1))
                    break
            m4 = re.search(r"\(\s*([\d.,]+)\s*%", riga)
            if m4:
                d["iva_percento"] = num_en(m4.group(1))
            break
    d["iva"] = iva_usd

    m = re.search(r"Amount due\s+\$?([\d,.]+)\s*([A-Z]{3})?", t)
    if m:
        d["totale"] = num_en(m.group(1))
        if m.group(2):
            d["valuta"] = m.group(2)
    else:
        tot = re.findall(r"^\s*Total\s+\$([\d,.]+)\s*$", t, re.M)
        d["totale"] = num_en(tot[-1]) if tot else None

    # OpenAI riporta l'IVA anche in EUR: da li' ricavo il cambio effettivo
    m = re.search(r"\(\s*€\s*([\d,.]+)\s*\)", t)
    if m and iva_usd:
        iva_eur = num_en(m.group(1))
        if iva_eur and iva_usd:
            d["cambio"] = round(iva_eur / iva_usd, 6)
            d["cambio_fonte"] = "ricavato dall'IVA in EUR riportata in fattura"

    m = re.search(r"^\s*(.+?)\s{2,}\d+\s{2,}\$", t, re.M)
    if m:
        d["descrizione"] = m.group(1).strip()

    return d


def parse_generico(t: str) -> dict:
    """Fallback per fornitori non ancora mappati: tenta i pattern piu' comuni."""
    d: dict = {"valuta": "EUR", "avvisi": ["Parser generico: controlla i valori a mano."]}

    m = (re.search(r"(?:Invoice number|Numero fattura|Fattura n[.°]?)\s*[:#]?\s*(\S+)", t, re.I))
    d["numero"] = m.group(1) if m else None

    m = re.search(r"(\d{2}/\d{2}/\d{4})", t)
    if m:
        d["data"] = iso_da_it(m.group(1))
    else:
        m = re.search(r"([A-Za-z]+\s+\d{1,2},\s*\d{4})", t)
        d["data"] = iso_da_en(m.group(1)) if m else None

    if "$" in t and "€" not in t:
        d["valuta"] = "USD"

    m = re.search(r"(?:Total|Totale)[^\n\d]*([\d.,]+)", t, re.I)
    if m:
        d["totale"] = num_en(m.group(1)) if d["valuta"] == "USD" else num_it(m.group(1))

    return d


PARSER = {"ionos": parse_ionos, "openai": parse_openai, "generico": parse_generico}


# --------------------------------------------------------------------------
# import
# --------------------------------------------------------------------------

def carica_config() -> dict:
    return json.loads(CONFIG_PATH.read_text(encoding="utf-8"))


def carica_spese() -> dict:
    if SPESE_JSON.exists():
        return json.loads(SPESE_JSON.read_text(encoding="utf-8"))
    return {"fatture": [], "ultimo_import": None}


def importa(cfg: dict, stato: dict, reimporta: bool) -> tuple[int, list[str]]:
    gia_viste = set() if reimporta else {f["hash"] for f in stato["fatture"]}
    if reimporta:
        stato["fatture"] = []

    nuove, problemi = 0, []

    for pdf in sorted(BASE.rglob("*.pdf")):
        if any(parte in CARTELLE_IGNORATE for parte in pdf.relative_to(BASE).parts):
            continue

        h = sha_file(pdf)
        if h in gia_viste:
            continue

        rel = pdf.relative_to(BASE)
        chiave = rel.parts[0].lower() if len(rel.parts) > 1 else "sconosciuto"
        info_forn = cfg["fornitori"].get(chiave, {})
        nome_parser = info_forn.get("parser", "generico")

        try:
            testo = pdf_testo(pdf)
        except RuntimeError as e:
            problemi.append(f"{rel}: {e}")
            continue

        if not testo.strip():
            problemi.append(f"{rel}: PDF senza testo estraibile (probabile scansione, serve OCR).")
            continue

        d = PARSER[nome_parser](testo)

        valuta = d.get("valuta") or info_forn.get("valuta", "EUR")
        cambio = d.get("cambio")
        if valuta != "EUR" and not cambio:
            cambio = cfg.get("cambio_fallback_usd_eur", 1.0)
            d.setdefault("avvisi", []).append(
                f"Cambio non presente in fattura: usato il fallback {cambio} da config.json."
            )
            d["cambio_fonte"] = "fallback da config.json"
        if valuta == "EUR":
            cambio = 1.0

        def in_eur(v):
            return round(v * cambio, 2) if v is not None else None

        # coerenza netto + IVA = totale
        netto, iva, totale = d.get("netto"), d.get("iva"), d.get("totale")
        if netto is not None and iva is not None and totale is not None:
            if abs((netto + iva) - totale) > 0.02:
                d.setdefault("avvisi", []).append(
                    f"Netto {netto} + IVA {iva} != Totale {totale}: verifica manualmente."
                )
        if totale is None:
            d.setdefault("avvisi", []).append("Totale non riconosciuto: inseriscilo a mano.")

        record = {
            "hash": h,
            "file": str(rel).replace("\\", "/"),
            "fornitore": chiave,
            "fornitore_nome": info_forn.get("nome", chiave),
            "categoria": info_forn.get("categoria", "Altro"),
            "numero": d.get("numero"),
            "data": d.get("data"),
            "descrizione": d.get("descrizione"),
            "contratto": d.get("contratto"),
            "periodo_da": d.get("periodo_da"),
            "periodo_a": d.get("periodo_a"),
            "sconto_a": d.get("sconto_a"),
            "valuta": valuta,
            "cambio": cambio,
            "cambio_fonte": d.get("cambio_fonte", "n/a" if valuta == "EUR" else "fallback"),
            "netto": netto,
            "iva": iva,
            "iva_percento": d.get("iva_percento"),
            "totale": totale,
            "netto_eur": in_eur(netto),
            "iva_eur": in_eur(iva),
            "totale_eur": in_eur(totale),
            "voci": d.get("voci", []),
            "avvisi": d.get("avvisi", []),
            "origine": "pdf",
        }
        stato["fatture"].append(record)
        nuove += 1
        if record["avvisi"]:
            problemi.extend(f"{rel}: {a}" for a in record["avvisi"])

    # spese inserite a mano
    stato["fatture"] = [f for f in stato["fatture"] if f.get("origine") != "manuale"]
    if MANUALI_CSV.exists():
        with MANUALI_CSV.open(encoding="utf-8-sig", newline="") as f:
            for i, riga in enumerate(csv.DictReader(f), start=2):
                if not (riga.get("data") or "").strip():
                    continue
                try:
                    netto = float((riga.get("netto") or "0").replace(",", "."))
                    iva = float((riga.get("iva") or "0").replace(",", "."))
                    cambio = float((riga.get("cambio") or "1").replace(",", ".")) or 1.0
                except ValueError:
                    problemi.append(f"spese_manuali.csv riga {i}: importi non numerici.")
                    continue
                stato["fatture"].append({
                    "hash": f"manuale-{i}-{riga.get('data')}",
                    "file": "spese_manuali.csv",
                    "fornitore": (riga.get("fornitore") or "altro").strip().lower(),
                    "fornitore_nome": (riga.get("fornitore") or "Altro").strip(),
                    "categoria": (riga.get("categoria") or "Altro").strip(),
                    "numero": (riga.get("numero") or "").strip() or None,
                    "data": riga["data"].strip(),
                    "descrizione": (riga.get("descrizione") or "").strip(),
                    "contratto": None, "periodo_da": None, "periodo_a": None, "sconto_a": None,
                    "valuta": (riga.get("valuta") or "EUR").strip().upper(),
                    "cambio": cambio, "cambio_fonte": "manuale",
                    "netto": netto, "iva": iva,
                    "iva_percento": round(iva / netto * 100, 1) if netto else None,
                    "totale": round(netto + iva, 2),
                    "netto_eur": round(netto * cambio, 2),
                    "iva_eur": round(iva * cambio, 2),
                    "totale_eur": round((netto + iva) * cambio, 2),
                    "voci": [], "avvisi": [], "origine": "manuale",
                })

    stato["fatture"].sort(key=lambda f: (f.get("data") or "", f.get("fornitore") or ""))
    stato["ultimo_import"] = datetime.now().isoformat(timespec="seconds")
    return nuove, problemi


# --------------------------------------------------------------------------
# metriche dal VPS
# --------------------------------------------------------------------------

def scarica_metriche(cfg: dict) -> tuple[bool, str]:
    """Esegue metriche.sql sul Postgres di produzione via SSH.

    Non solleva mai: se qualcosa va storto lo segnala e lascia intatto
    l'eventuale metriche.json precedente.
    """
    vps = cfg.get("vps") or {}
    host = os.environ.get("VPS_HOST") or vps.get("host")
    cartella = os.environ.get("VPS_DIR") or vps.get("cartella", "/opt/trainmind")
    servizio = os.environ.get("DB_SERVICE") or vps.get("servizio_db", "db")
    utente = os.environ.get("DB_USER") or vps.get("db_utente", "trainmind")
    database = os.environ.get("DB_NAME") or vps.get("db_nome", "trainmind")

    if not host:
        return False, "host del VPS non configurato (config.json -> vps.host)."
    if not shutil.which("ssh"):
        return False, "comando 'ssh' non trovato su questa macchina."
    if not SQL_METRICHE.exists():
        return False, f"{SQL_METRICHE.name} non trovato."

    remoto = (
        f"cd {cartella} && docker compose exec -T {servizio} "
        f"psql -U {utente} -d {database} -At -f -"
    )
    try:
        res = subprocess.run(
            ["ssh", "-o", "BatchMode=yes", "-o", "ConnectTimeout=15", host, remoto],
            input=SQL_METRICHE.read_text(encoding="utf-8"),
            capture_output=True, text=True, encoding="utf-8", errors="replace",
            timeout=90,
        )
    except subprocess.TimeoutExpired:
        return False, "il VPS non ha risposto entro 90 secondi."
    except OSError as e:
        return False, f"impossibile avviare ssh: {e}"

    if res.returncode != 0:
        err = (res.stderr or "").strip().splitlines()
        dettaglio = err[-1] if err else f"exit code {res.returncode}"
        return False, f"connessione o query fallita: {dettaglio}"

    grezzo = (res.stdout or "").strip()
    if not grezzo:
        return False, "la query non ha restituito nulla."

    # psql -At stampa una riga sola, ma tolgo eventuali righe di rumore
    riga = next((r for r in reversed(grezzo.splitlines()) if r.strip().startswith("{")), None)
    if not riga:
        return False, f"output non riconosciuto: {grezzo[:120]}"

    try:
        dati = json.loads(riga)
    except json.JSONDecodeError as e:
        return False, f"JSON non valido dal DB: {e}"

    DATI.mkdir(exist_ok=True)
    METRICHE_JSON.write_text(json.dumps(dati, ensure_ascii=False, indent=2), encoding="utf-8")
    return True, (
        f"{dati.get('organizzazioni_totali', '?')} organizzazioni, "
        f"{dati.get('utenti_attivi', '?')} utenti attivi, "
        f"{dati.get('costo_ai_usd_mese_corrente', 0)} USD di AI questo mese"
    )


# --------------------------------------------------------------------------
# analisi
# --------------------------------------------------------------------------

def analizza(cfg: dict, stato: dict) -> dict:
    fatture = stato["fatture"]
    oggi = date.today()

    per_mese: dict[str, dict] = defaultdict(lambda: {"netto": 0.0, "iva": 0.0, "totale": 0.0})
    per_fornitore: dict[str, float] = defaultdict(float)
    per_categoria: dict[str, float] = defaultdict(float)
    per_anno: dict[str, dict] = defaultdict(lambda: {"netto": 0.0, "iva": 0.0, "totale": 0.0})

    for f in fatture:
        if not f.get("data"):
            continue
        mese, anno = f["data"][:7], f["data"][:4]
        for chiave, campo in (("netto", "netto_eur"), ("iva", "iva_eur"), ("totale", "totale_eur")):
            v = f.get(campo) or 0.0
            per_mese[mese][chiave] += v
            per_anno[anno][chiave] += v
        per_fornitore[f.get("fornitore_nome") or "?"] += f.get("totale_eur") or 0.0
        per_categoria[f.get("categoria") or "Altro"] += f.get("totale_eur") or 0.0

    for d in list(per_mese.values()) + list(per_anno.values()):
        for k in d:
            d[k] = round(d[k], 2)

    # --- costo ricorrente previsto ---
    ricorrente_mensile = 0.0
    for c in cfg["contratti"]:
        att = c.get("costo_netto_attuale")
        if att is None:
            continue
        lordo = att * (1 + (c.get("iva_percento") or 0) / 100)
        if c["ricorrenza"] == "mensile":
            ricorrente_mensile += lordo
        elif c["ricorrenza"] == "annuale":
            ricorrente_mensile += lordo / 12

    ricorrente_mensile_pieno = 0.0
    for c in cfg["contratti"]:
        pieno = c.get("costo_netto_pieno")
        if pieno is None:
            continue
        lordo = pieno * (1 + (c.get("iva_percento") or 0) / 100)
        if c["ricorrenza"] == "mensile":
            ricorrente_mensile_pieno += lordo
        elif c["ricorrenza"] == "annuale":
            ricorrente_mensile_pieno += lordo / 12

    # --- scadenze ---
    scadenze = []
    for c in cfg["contratti"]:
        for campo, tipo in (("sconto_fino_al", "Fine sconto"), ("scade_il", "Rinnovo")):
            quando = c.get(campo)
            if not quando:
                continue
            d_scad = date.fromisoformat(quando)
            giorni = (d_scad - oggi).days
            att = c.get("costo_netto_attuale")
            pieno = c.get("costo_netto_pieno")
            delta = None
            if att is not None and pieno is not None and pieno != att:
                iva_f = 1 + (c.get("iva_percento") or 0) / 100
                delta = round((pieno - att) * iva_f, 2)
            scadenze.append({
                "contratto": c["descrizione"],
                "tipo": tipo,
                "data": quando,
                "giorni": giorni,
                "aumento_eur": delta,
                "periodicita": c["ricorrenza"],
                "nota": c.get("nota_sconto"),
                "urgenza": "scaduto" if giorni < 0 else ("alta" if giorni <= 30 else ("media" if giorni <= 90 else "bassa")),
            })
    # se sulla stessa data cadono piu' eventi dello stesso contratto, li unisco
    viste: dict[tuple, dict] = {}
    for s in sorted(scadenze, key=lambda x: x["data"]):
        k = (s["contratto"], s["data"])
        if k in viste:
            tipi = viste[k]["tipo"].split(" + ")
            if s["tipo"] not in tipi:
                viste[k]["tipo"] += " + " + s["tipo"]
            continue
        viste[k] = s
    scadenze = list(viste.values())

    # --- previsione 12 mesi ---
    previsione = []
    cursore = date(oggi.year, oggi.month, 1)
    for _ in range(12):
        mese_iso = cursore.isoformat()[:7]
        tot = 0.0
        for c in cfg["contratti"]:
            iva_f = 1 + (c.get("iva_percento") or 0) / 100
            att, pieno = c.get("costo_netto_attuale"), c.get("costo_netto_pieno")
            if att is None:
                continue
            fine_sconto = c.get("sconto_fino_al")
            costo = att
            if fine_sconto and cursore > date.fromisoformat(fine_sconto) and pieno is not None:
                costo = pieno
            if c["ricorrenza"] == "mensile":
                tot += costo * iva_f
            elif c["ricorrenza"] == "annuale":
                scad = c.get("scade_il")
                if scad and date.fromisoformat(scad).strftime("%Y-%m") == cursore.strftime("%Y-%m"):
                    tot += (pieno if pieno is not None else costo) * iva_f
        previsione.append({"mese": mese_iso, "previsto": round(tot, 2)})
        cursore = (cursore.replace(day=28) + timedelta(days=8)).replace(day=1)

    # --- IVA per trimestre ---
    iva_trim: dict[str, float] = defaultdict(float)
    for f in fatture:
        if not f.get("data"):
            continue
        d = date.fromisoformat(f["data"])
        iva_trim[f"{d.year} T{(d.month - 1) // 3 + 1}"] += f.get("iva_eur") or 0.0
    iva_trim = {k: round(v, 2) for k, v in sorted(iva_trim.items())}

    # --- metriche utenti dal VPS ---
    metriche = json.loads(METRICHE_JSON.read_text(encoding="utf-8")) if METRICHE_JSON.exists() else None

    mese_corr = oggi.isoformat()[:7]
    speso_mese = per_mese.get(mese_corr, {}).get("totale", 0.0)
    base_costo = max(speso_mese, round(ricorrente_mensile, 2))

    breakeven = None
    if metriche:
        org = metriche.get("organizzazioni_totali") or 0
        org_pag = metriche.get("organizzazioni_paganti") or 0
        utenti = metriche.get("utenti_attivi") or 0
        atleti = metriche.get("atleti_attivi") or 0
        ai_mese = metriche.get("costo_ai_usd_mese_corrente") or 0.0
        cambio = cfg.get("cambio_fallback_usd_eur", 0.877)
        breakeven = {
            "aggiornato_al": metriche.get("generato_il"),
            "organizzazioni_totali": org,
            "organizzazioni_paganti": org_pag,
            "utenti_attivi": utenti,
            "atleti_attivi": atleti,
            "costo_mensile_eur": round(base_costo, 2),
            "costo_per_organizzazione": round(base_costo / org, 2) if org else None,
            "costo_per_utente": round(base_costo / utenti, 2) if utenti else None,
            "costo_per_atleta": round(base_costo / atleti, 2) if atleti else None,
            "costo_ai_eur_mese": round(ai_mese * cambio, 2),
            "quota_ai_percento": round(ai_mese * cambio / base_costo * 100, 1) if base_costo else None,
            "ricavo_mensile_eur": metriche.get("ricavo_mensile_eur"),
        }
        ric = metriche.get("ricavo_mensile_eur")
        if ric is not None:
            breakeven["margine_eur"] = round(ric - base_costo, 2)
            breakeven["in_pareggio"] = ric >= base_costo

    return {
        "generato_il": datetime.now().isoformat(timespec="seconds"),
        "n_fatture": len(fatture),
        "totale_storico": round(sum(f.get("totale_eur") or 0.0 for f in fatture), 2),
        "iva_storica": round(sum(f.get("iva_eur") or 0.0 for f in fatture), 2),
        "per_mese": dict(sorted(per_mese.items())),
        "per_anno": dict(sorted(per_anno.items())),
        "per_fornitore": {k: round(v, 2) for k, v in sorted(per_fornitore.items(), key=lambda x: -x[1])},
        "per_categoria": {k: round(v, 2) for k, v in sorted(per_categoria.items(), key=lambda x: -x[1])},
        "ricorrente_mensile": round(ricorrente_mensile, 2),
        "ricorrente_annuale": round(ricorrente_mensile * 12, 2),
        "ricorrente_mensile_senza_sconti": round(ricorrente_mensile_pieno, 2),
        "scadenze": scadenze,
        "previsione_12m": previsione,
        "previsione_totale_12m": round(sum(p["previsto"] for p in previsione), 2),
        "iva_per_trimestre": iva_trim,
        "breakeven": breakeven,
        "metriche_presenti": metriche is not None,
    }


# --------------------------------------------------------------------------
# Excel
# --------------------------------------------------------------------------

def scrivi_excel(cfg: dict, stato: dict, an: dict) -> None:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    BLU = "1F3A5F"
    testa = Font(bold=True, color="FFFFFF", size=10)
    riemp = PatternFill("solid", fgColor=BLU)
    EURO = '#,##0.00\\ "€"'

    wb = Workbook()

    def foglio(titolo, intestazioni, righe, larghezze=None, formati=None):
        ws = wb.create_sheet(titolo)
        ws.append(intestazioni)
        for c in ws[1]:
            c.font, c.fill = testa, riemp
            c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        for r in righe:
            ws.append(r)
        for i, larg in enumerate(larghezze or [], start=1):
            ws.column_dimensions[get_column_letter(i)].width = larg
        for col, fmt in (formati or {}).items():
            for riga in ws.iter_rows(min_row=2, min_col=col, max_col=col):
                for c in riga:
                    c.number_format = fmt
        ws.freeze_panes = "A2"
        return ws

    # 1. Registro
    foglio(
        "Registro",
        ["Data", "Fornitore", "Categoria", "Descrizione", "N. fattura", "Valuta",
         "Netto", "IVA", "Totale", "Cambio", "Netto EUR", "IVA EUR", "Totale EUR",
         "Periodo da", "Periodo a", "File", "Note"],
        [[
            f.get("data"), f.get("fornitore_nome"), f.get("categoria"), f.get("descrizione"),
            f.get("numero"), f.get("valuta"), f.get("netto"), f.get("iva"), f.get("totale"),
            f.get("cambio"), f.get("netto_eur"), f.get("iva_eur"), f.get("totale_eur"),
            f.get("periodo_da"), f.get("periodo_a"), f.get("file"),
            " | ".join(f.get("avvisi") or []),
        ] for f in stato["fatture"]],
        [12, 20, 18, 34, 18, 8, 11, 10, 11, 9, 12, 11, 12, 12, 12, 44, 40],
        {7: EURO, 8: EURO, 9: EURO, 11: EURO, 12: EURO, 13: EURO},
    )

    # 2. Per mese
    foglio(
        "Per mese",
        ["Mese", "Netto EUR", "IVA EUR", "Totale EUR"],
        [[m, v["netto"], v["iva"], v["totale"]] for m, v in an["per_mese"].items()],
        [14, 14, 14, 14], {2: EURO, 3: EURO, 4: EURO},
    )

    # 3. Contratti e rinnovi
    foglio(
        "Contratti e rinnovi",
        ["Contratto", "Fornitore", "Ricorrenza", "Costo attuale (netto)",
         "Costo pieno (netto)", "IVA %", "Lordo attuale", "Scadenza sconto",
         "Scadenza rinnovo", "Nota"],
        [[
            c["descrizione"], cfg["fornitori"].get(c["fornitore"], {}).get("nome", c["fornitore"]),
            c["ricorrenza"], c.get("costo_netto_attuale"), c.get("costo_netto_pieno"),
            c.get("iva_percento"),
            round((c["costo_netto_attuale"] or 0) * (1 + (c.get("iva_percento") or 0) / 100), 2)
            if c.get("costo_netto_attuale") is not None else None,
            c.get("sconto_fino_al"), c.get("scade_il"), c.get("nota_sconto"),
        ] for c in cfg["contratti"]],
        [38, 18, 13, 20, 20, 8, 15, 17, 17, 80],
        {4: EURO, 5: EURO, 7: EURO},
    )

    # 4. Scadenze
    foglio(
        "Scadenze",
        ["Data", "Giorni", "Urgenza", "Tipo", "Contratto", "Impatto EUR", "Nota"],
        [[s["data"], s["giorni"], s["urgenza"].upper(), s["tipo"], s["contratto"],
          s.get("aumento_eur"), s.get("nota")] for s in an["scadenze"]],
        [12, 9, 11, 14, 38, 14, 90], {6: EURO},
    )

    # 5. Previsione 12 mesi
    foglio(
        "Previsione 12 mesi",
        ["Mese", "Costo previsto EUR (IVA incl.)"],
        [[p["mese"], p["previsto"]] for p in an["previsione_12m"]]
        + [["TOTALE", an["previsione_totale_12m"]]],
        [14, 28], {2: EURO},
    )

    # 6. IVA
    foglio(
        "IVA",
        ["Periodo", "IVA EUR"],
        [[k, v] for k, v in an["iva_per_trimestre"].items()]
        + [["TOTALE", an["iva_storica"]]],
        [16, 16], {2: EURO},
    )

    # 7. Break-even
    be = an.get("breakeven")
    if be:
        righe = [
            ["Aggiornato al", be.get("aggiornato_al")],
            ["Organizzazioni totali", be.get("organizzazioni_totali")],
            ["Organizzazioni paganti", be.get("organizzazioni_paganti")],
            ["Utenti attivi", be.get("utenti_attivi")],
            ["Atleti attivi", be.get("atleti_attivi")],
            ["Costo mensile EUR", be.get("costo_mensile_eur")],
            ["Costo per organizzazione", be.get("costo_per_organizzazione")],
            ["Costo per utente", be.get("costo_per_utente")],
            ["Costo per atleta", be.get("costo_per_atleta")],
            ["Costo AI del mese (EUR)", be.get("costo_ai_eur_mese")],
            ["Quota AI sul totale (%)", be.get("quota_ai_percento")],
            ["Ricavo mensile EUR", be.get("ricavo_mensile_eur")],
            ["Margine EUR", be.get("margine_eur")],
            ["In pareggio", "SI" if be.get("in_pareggio") else "NO"],
        ]
    else:
        righe = [
            ["Metriche non ancora scaricate dal VPS.", ""],
            ["Come fare", "Lancia aggiorna.bat (o: bash aggiorna.sh) dalla cartella Fatture."],
            ["Requisito", "Chiave SSH configurata per il server."],
            ["Impostazioni connessione", "_sistema/config.json, sezione \"vps\"."],
        ]
    foglio("Break-even", ["Voce", "Valore"], righe, [34, 62])

    # 8. Voci di dettaglio
    dett = [[f.get("data"), f.get("fornitore_nome"), f.get("numero"), v.get("n"),
             v.get("descrizione"), v.get("netto"), v.get("iva_percento")]
            for f in stato["fatture"] for v in (f.get("voci") or [])]
    foglio("Voci fattura", ["Data", "Fornitore", "N. fattura", "Voce", "Descrizione",
                            "Netto", "IVA %"], dett, [12, 18, 18, 8, 46, 12, 9], {6: EURO})

    # copertina
    ws = wb["Sheet"]
    ws.title = "Riepilogo"
    ws.column_dimensions["A"].width = 40
    ws.column_dimensions["B"].width = 26
    ws["A1"] = "TrainMind — Registro spese"
    ws["A1"].font = Font(bold=True, size=16, color=BLU)
    riepilogo = [
        ("", ""),
        ("Generato il", an["generato_il"].replace("T", " ")),
        ("Fatture registrate", an["n_fatture"]),
        ("", ""),
        ("Totale speso (storico)", an["totale_storico"]),
        ("di cui IVA", an["iva_storica"]),
        ("", ""),
        ("Costo ricorrente / mese (oggi)", an["ricorrente_mensile"]),
        ("Costo ricorrente / anno (oggi)", an["ricorrente_annuale"]),
        ("Costo ricorrente / mese senza sconti", an["ricorrente_mensile_senza_sconti"]),
        ("Previsione prossimi 12 mesi", an["previsione_totale_12m"]),
    ]
    for i, (k, v) in enumerate(riepilogo, start=2):
        ws.cell(row=i, column=1, value=k).font = Font(bold=bool(k))
        c = ws.cell(row=i, column=2, value=v)
        if isinstance(v, (int, float)) and k not in ("Fatture registrate",):
            c.number_format = EURO
    prossima = next((s for s in an["scadenze"] if s["giorni"] >= 0), None)
    if prossima:
        r = len(riepilogo) + 3
        ws.cell(row=r, column=1, value="Prossima scadenza").font = Font(bold=True, color="B00020")
        ws.cell(row=r, column=2, value=f"{prossima['data']} — {prossima['contratto']} ({prossima['giorni']} gg)")

    wb.move_sheet("Riepilogo", offset=-wb.index(wb["Riepilogo"]))
    wb.save(XLSX_OUT)


# --------------------------------------------------------------------------
# Dashboard HTML
# --------------------------------------------------------------------------

def scrivi_html(cfg: dict, stato: dict, an: dict) -> None:
    dati = json.dumps(
        {"analisi": an, "fatture": stato["fatture"], "contratti": cfg["contratti"]},
        ensure_ascii=False,
    )
    html = HTML_TEMPLATE.replace("/*__DATI__*/null", dati)
    HTML_OUT.write_text(html, encoding="utf-8")


HTML_TEMPLATE = r"""<!DOCTYPE html>
<html lang="it">
<head>
<meta charset="utf-8">
<meta name="viewport" content="width=device-width, initial-scale=1">
<title>TrainMind — Spese infrastruttura</title>
<script src="https://cdn.jsdelivr.net/npm/chart.js@4.4.1/dist/chart.umd.min.js"></script>
<style>
  :root{--bg:#0f1720;--card:#182430;--bd:#24384a;--tx:#e8eef4;--mut:#8fa6ba;
        --acc:#4da3ff;--ok:#3ecf8e;--warn:#ffb020;--bad:#ff5c5c;}
  *{box-sizing:border-box}
  body{margin:0;background:var(--bg);color:var(--tx);
       font-family:system-ui,-apple-system,"Segoe UI",Roboto,sans-serif;padding:28px 20px 60px}
  .wrap{max-width:1180px;margin:0 auto}
  h1{margin:0 0 4px;font-size:26px}
  .sub{color:var(--mut);font-size:13px;margin-bottom:26px}
  h2{font-size:16px;margin:34px 0 12px;color:var(--acc);
     text-transform:uppercase;letter-spacing:.7px}
  .kpis{display:grid;grid-template-columns:repeat(auto-fit,minmax(190px,1fr));gap:14px}
  .kpi{background:var(--card);border:1px solid var(--bd);border-radius:12px;padding:16px 18px}
  .kpi .lab{color:var(--mut);font-size:11.5px;text-transform:uppercase;letter-spacing:.6px}
  .kpi .val{font-size:25px;font-weight:650;margin-top:6px}
  .kpi .note{color:var(--mut);font-size:11.5px;margin-top:5px;line-height:1.45}
  .card{background:var(--card);border:1px solid var(--bd);border-radius:12px;padding:18px}
  .grid2{display:grid;grid-template-columns:1fr 1fr;gap:16px}
  @media(max-width:820px){.grid2{grid-template-columns:1fr}}
  table{width:100%;border-collapse:collapse;font-size:13px}
  th{text-align:left;color:var(--mut);font-weight:600;font-size:11px;
     text-transform:uppercase;letter-spacing:.5px;padding:9px 10px;border-bottom:1px solid var(--bd)}
  td{padding:9px 10px;border-bottom:1px solid rgba(255,255,255,.05)}
  tr:last-child td{border-bottom:none}
  .num{text-align:right;font-variant-numeric:tabular-nums}
  .pill{display:inline-block;padding:2px 9px;border-radius:99px;font-size:11px;font-weight:600}
  .p-alta{background:rgba(255,92,92,.16);color:var(--bad)}
  .p-media{background:rgba(255,176,32,.16);color:var(--warn)}
  .p-bassa{background:rgba(62,207,142,.14);color:var(--ok)}
  .p-scaduto{background:rgba(143,166,186,.16);color:var(--mut)}
  .empty{color:var(--mut);font-size:13px;padding:14px 4px}
  .warn{background:rgba(255,176,32,.1);border:1px solid rgba(255,176,32,.35);
        border-radius:10px;padding:13px 16px;font-size:13px;color:#ffd9a0}
  canvas{max-height:270px}
  /* barra di aggiornamento */
  .top{display:flex;align-items:flex-start;justify-content:space-between;gap:18px;flex-wrap:wrap}
  .azioni{display:flex;align-items:center;gap:10px;flex-wrap:wrap}
  button{font:inherit;font-size:13px;font-weight:600;cursor:pointer;border-radius:9px;
         padding:9px 16px;border:1px solid var(--bd);background:var(--card);color:var(--tx)}
  button:hover:not(:disabled){border-color:var(--acc)}
  button:disabled{opacity:.5;cursor:not-allowed}
  button.primario{background:var(--acc);border-color:var(--acc);color:#06121f}
  button.primario:hover:not(:disabled){filter:brightness(1.1)}
  .chk{display:flex;align-items:center;gap:6px;font-size:12px;color:var(--mut);cursor:pointer}
  #esito{margin:14px 0 0;border-radius:10px;font-size:12.5px;display:none}
  #esito.on{display:block}
  #esito .head{padding:11px 15px;font-weight:600}
  #esito.ok .head{background:rgba(62,207,142,.14);color:var(--ok)}
  #esito.ko .head{background:rgba(255,92,92,.14);color:var(--bad)}
  #esito.run .head{background:rgba(77,163,255,.14);color:var(--acc)}
  #esito pre{margin:0;padding:13px 15px;background:var(--card);border:1px solid var(--bd);
             border-top:none;border-radius:0 0 10px 10px;white-space:pre-wrap;
             font-family:ui-monospace,"Cascadia Code",Consolas,monospace;
             font-size:12px;line-height:1.55;color:var(--mut);max-height:290px;overflow:auto}
  .hint{background:var(--card);border:1px solid var(--bd);border-radius:10px;
        padding:12px 15px;font-size:12.5px;color:var(--mut);line-height:1.6;max-width:430px}
  .hint code{background:rgba(255,255,255,.07);padding:2px 6px;border-radius:5px;
             font-family:ui-monospace,Consolas,monospace;color:var(--tx)}
</style>
</head>
<body>
<div class="wrap">
  <div class="top">
    <div>
      <h1>TrainMind — Spese infrastruttura</h1>
      <div class="sub" id="sub"></div>
    </div>
    <div id="barra"></div>
  </div>
  <div id="esito"><div class="head"></div><pre></pre></div>
  <div id="app"></div>
</div>
<script>
const D = /*__DATI__*/null;
const A = D.analisi;
const eur = n => (n===null||n===undefined) ? "—" :
  new Intl.NumberFormat("it-IT",{style:"currency",currency:"EUR"}).format(n);
const mesiIt=["","gen","feb","mar","apr","mag","giu","lug","ago","set","ott","nov","dic"];
const mLab = m => { const [a,b]=m.split("-"); return mesiIt[+b]+" "+a.slice(2); };
const esc = s => String(s??"").replace(/[<>&]/g,c=>({"<":"&lt;",">":"&gt;","&":"&amp;"}[c]));

document.getElementById("sub").textContent =
  "Aggiornato il " + A.generato_il.replace("T"," ") + " · " + A.n_fatture + " fatture registrate";

/* ---- barra di aggiornamento -------------------------------------------
   Il bottone puo' funzionare solo se la pagina arriva dal server locale.
   Aperta con doppio clic (file://) il browser vieta di lanciare programmi,
   quindi in quel caso mostro il comando da copiare invece di un bottone finto. */
const vivo = location.protocol === "http:" || location.protocol === "https:";
const barra = document.getElementById("barra");
const esito = document.getElementById("esito");

if (vivo) {
  barra.innerHTML = `<div class="azioni">
      <label class="chk"><input type="checkbox" id="chkVps" checked> includi metriche dal VPS</label>
      <button id="btnAgg" class="primario">Aggiorna ora</button>
    </div>`;
  document.getElementById("btnAgg").addEventListener("click", aggiorna);
} else {
  barra.innerHTML = `<div class="hint">
    Per aggiornare con un bottone lancia <code>aggiorna.bat --dashboard</code>
    (o <code>bash aggiorna.sh --dashboard</code>): apre questa stessa pagina da un
    server locale, dove il bottone funziona.<br><br>
    Aperta con un doppio clic il browser non puo' eseguire programmi sul computer,
    quindi qui il bottone sarebbe solo decorativo.</div>`;
}

function mostra(stato, titolo, testo) {
  esito.className = "on " + stato;
  esito.querySelector(".head").textContent = titolo;
  const pre = esito.querySelector("pre");
  pre.textContent = testo || "";
  pre.style.display = testo ? "" : "none";
}

async function aggiorna() {
  const btn = document.getElementById("btnAgg");
  const conVps = document.getElementById("chkVps").checked;
  btn.disabled = true;
  const testoOrig = btn.textContent;
  btn.textContent = "Aggiornamento…";
  mostra("run", conVps ? "In corso — collegamento al VPS, lettura fatture, report…"
                       : "In corso — lettura fatture, report…", "");
  try {
    const r = await fetch("/api/aggiorna?metriche=" + (conVps ? "1" : "0"), { method: "POST" });
    const d = await r.json();
    if (d.ok) {
      mostra("ok", "Fatto — ricarico la pagina…", d.output);
      setTimeout(() => location.reload(), 1400);
    } else {
      mostra("ko", "L'aggiornamento non e' andato a buon fine", d.output || d.errore || "");
      btn.disabled = false; btn.textContent = testoOrig;
    }
  } catch (e) {
    mostra("ko", "Server locale non raggiungibile",
      "La finestra del terminale che serve la dashboard e' stata chiusa?\n\n" + e);
    btn.disabled = false; btn.textContent = testoOrig;
  }
}

const be = A.breakeven;
const prossima = A.scadenze.find(s => s.giorni >= 0);

let h = "";

/* KPI */
h += `<div class="kpis">
  <div class="kpi"><div class="lab">Costo ricorrente / mese</div>
    <div class="val">${eur(A.ricorrente_mensile)}</div>
    <div class="note">IVA inclusa, ai prezzi scontati di oggi</div></div>
  <div class="kpi"><div class="lab">Su base annua</div>
    <div class="val">${eur(A.ricorrente_annuale)}</div>
    <div class="note">Senza sconti sarebbe ${eur(A.ricorrente_mensile_senza_sconti*12)}</div></div>
  <div class="kpi"><div class="lab">Previsione 12 mesi</div>
    <div class="val">${eur(A.previsione_totale_12m)}</div>
    <div class="note">Tiene conto della fine degli sconti</div></div>
  <div class="kpi"><div class="lab">Totale speso finora</div>
    <div class="val">${eur(A.totale_storico)}</div>
    <div class="note">di cui ${eur(A.iva_storica)} di IVA</div></div>
</div>`;

/* Scadenze */
h += `<h2>Rinnovi e scadenze</h2><div class="card">`;
if (A.scadenze.length) {
  h += `<table><thead><tr><th>Data</th><th>Tra</th><th>Tipo</th><th>Contratto</th>
        <th class="num">Impatto / mese</th><th>Cosa succede</th></tr></thead><tbody>`;
  for (const s of A.scadenze) {
    h += `<tr>
      <td>${s.data}</td>
      <td><span class="pill p-${s.urgenza}">${s.giorni<0?"scaduto":s.giorni+" gg"}</span></td>
      <td>${esc(s.tipo)}</td><td>${esc(s.contratto)}</td>
      <td class="num">${s.aumento_eur?("+"+eur(s.aumento_eur)):"—"}</td>
      <td style="color:var(--mut);font-size:12px">${esc(s.nota||"")}</td></tr>`;
  }
  h += `</tbody></table>`;
} else { h += `<div class="empty">Nessuna scadenza configurata.</div>`; }
h += `</div>`;

/* Grafici */
h += `<h2>Andamento</h2><div class="grid2">
  <div class="card"><canvas id="cMese"></canvas></div>
  <div class="card"><canvas id="cPrev"></canvas></div>
</div>`;

h += `<div class="grid2" style="margin-top:16px">
  <div class="card"><canvas id="cForn"></canvas></div>
  <div class="card">
    <table><thead><tr><th>Categoria</th><th class="num">Totale</th></tr></thead><tbody>`;
for (const [k,v] of Object.entries(A.per_categoria))
  h += `<tr><td>${esc(k)}</td><td class="num">${eur(v)}</td></tr>`;
h += `</tbody></table></div></div>`;

/* Break-even */
h += `<h2>Costo per utente e break-even</h2><div class="card">`;
if (be) {
  h += `<div class="kpis">
    <div class="kpi"><div class="lab">Organizzazioni</div><div class="val">${be.organizzazioni_totali}</div>
      <div class="note">${be.organizzazioni_paganti} paganti</div></div>
    <div class="kpi"><div class="lab">Costo per organizzazione</div><div class="val">${eur(be.costo_per_organizzazione)}</div>
      <div class="note">al mese</div></div>
    <div class="kpi"><div class="lab">Costo per atleta</div><div class="val">${eur(be.costo_per_atleta)}</div>
      <div class="note">${be.atleti_attivi} atleti attivi</div></div>
    <div class="kpi"><div class="lab">Quota AI</div><div class="val">${be.quota_ai_percento ?? "—"}%</div>
      <div class="note">${eur(be.costo_ai_eur_mese)} di OpenAI questo mese</div></div>`;
  if (be.margine_eur !== undefined && be.margine_eur !== null) {
    h += `<div class="kpi"><div class="lab">Margine mensile</div>
      <div class="val" style="color:${be.in_pareggio?"var(--ok)":"var(--bad)"}">${eur(be.margine_eur)}</div>
      <div class="note">${be.in_pareggio?"Sei in pareggio":"Sotto il break-even"}</div></div>`;
  }
  h += `</div><div class="note" style="color:var(--mut);font-size:12px;margin-top:12px">
    Metriche lette dal DB di produzione, aggiornate al ${esc(be.aggiornato_al||"?")}.</div>`;
} else {
  h += `<div class="empty">Metriche utenti non ancora scaricate dal VPS.<br><br>
    ${vivo ? `Premi <b>Aggiorna ora</b> qui sopra con la casella
      "includi metriche dal VPS" spuntata.` :
      `Lancia <code>aggiorna.bat</code> dalla cartella Fatture.`}
    Serve una chiave SSH gia' configurata per il server; le impostazioni di
    connessione stanno in <code>_sistema/config.json</code>, sezione <code>vps</code>.</div>`;
}
h += `</div>`;

/* IVA */
h += `<h2>IVA per trimestre</h2><div class="card"><table>
  <thead><tr><th>Periodo</th><th class="num">IVA assolta</th></tr></thead><tbody>`;
const ivaT = Object.entries(A.iva_per_trimestre);
if (ivaT.length) for (const [k,v] of ivaT) h += `<tr><td>${k}</td><td class="num">${eur(v)}</td></tr>`;
else h += `<tr><td colspan="2" class="empty">Nessun dato.</td></tr>`;
h += `<tr><td><b>Totale</b></td><td class="num"><b>${eur(A.iva_storica)}</b></td></tr>
  </tbody></table></div>`;

/* Registro */
h += `<h2>Tutte le fatture</h2><div class="card"><table>
  <thead><tr><th>Data</th><th>Fornitore</th><th>Descrizione</th><th>N.</th>
  <th class="num">Netto</th><th class="num">IVA</th><th class="num">Totale €</th></tr></thead><tbody>`;
for (const f of [...D.fatture].reverse()) {
  h += `<tr><td>${f.data??"—"}</td><td>${esc(f.fornitore_nome)}</td>
    <td>${esc(f.descrizione||"")}</td><td style="font-size:11px;color:var(--mut)">${esc(f.numero||"")}</td>
    <td class="num">${eur(f.netto_eur)}</td><td class="num">${eur(f.iva_eur)}</td>
    <td class="num"><b>${eur(f.totale_eur)}</b></td></tr>`;
}
h += `</tbody></table></div>`;

document.getElementById("app").innerHTML = h;

/* render grafici */
if (window.Chart) {
  const gr = "#8fa6ba", gd = "rgba(255,255,255,.06)";
  Chart.defaults.color = gr;
  Chart.defaults.borderColor = gd;
  const mesi = Object.keys(A.per_mese);
  new Chart(document.getElementById("cMese"), {
    type: "bar",
    data: { labels: mesi.map(mLab), datasets: [
      { label:"Netto", data: mesi.map(m=>A.per_mese[m].netto), backgroundColor:"#4da3ff" },
      { label:"IVA",   data: mesi.map(m=>A.per_mese[m].iva),   backgroundColor:"#2b5f8f" }]},
    options:{responsive:true, plugins:{title:{display:true,text:"Speso per mese (EUR)"}},
      scales:{x:{stacked:true},y:{stacked:true}}}
  });
  new Chart(document.getElementById("cPrev"), {
    type: "line",
    data: { labels: A.previsione_12m.map(p=>mLab(p.mese)),
      datasets:[{label:"Previsto", data:A.previsione_12m.map(p=>p.previsto),
        borderColor:"#ffb020", backgroundColor:"rgba(255,176,32,.15)", fill:true, tension:.25}]},
    options:{responsive:true, plugins:{title:{display:true,text:"Previsione prossimi 12 mesi (EUR, IVA incl.)"}}}
  });
  const fk = Object.keys(A.per_fornitore);
  new Chart(document.getElementById("cForn"), {
    type: "doughnut",
    data:{labels:fk, datasets:[{data:fk.map(k=>A.per_fornitore[k]),
      backgroundColor:["#4da3ff","#3ecf8e","#ffb020","#ff5c5c","#a78bfa"]}]},
    options:{responsive:true, plugins:{title:{display:true,text:"Ripartizione per fornitore"},
      legend:{position:"bottom"}}}
  });
}
</script>
</body>
</html>
"""


# --------------------------------------------------------------------------
# main
# --------------------------------------------------------------------------

def main() -> int:
    ap = argparse.ArgumentParser(description="Importa fatture e rigenera i report spese TrainMind.")
    ap.add_argument("--solo-report", action="store_true", help="non rilegge i PDF")
    ap.add_argument("--reimporta", action="store_true", help="rilegge tutti i PDF da zero")
    ap.add_argument("--senza-metriche", action="store_true",
                    help="salta il collegamento SSH al VPS per le metriche utenti")
    ap.add_argument("--dashboard", action="store_true",
                    help="al termine apre la dashboard da un server locale, con il bottone Aggiorna attivo")
    ap.add_argument("--porta", type=int, default=8765, help="porta del server (default 8765)")
    args = ap.parse_args()

    DATI.mkdir(exist_ok=True)
    cfg = carica_config()
    stato = carica_spese()

    print("=== Spese TrainMind ===\n")

    # 1) metriche dal VPS (non bloccante: se fallisce si prosegue)
    if args.senza_metriche:
        print("[1/3] Metriche VPS: saltate (--senza-metriche)")
    else:
        print("[1/3] Metriche dal VPS...", end=" ", flush=True)
        successo, msg = scarica_metriche(cfg)
        print(("OK — " if successo else "non disponibili — ") + msg)
        if not successo:
            print("      I report useranno le metriche precedenti, se presenti.")

    # 2) import fatture
    problemi: list[str] = []
    if args.solo_report:
        print(f"[2/3] Import PDF: saltato (--solo-report)")
    else:
        print("[2/3] Lettura fatture PDF...", end=" ", flush=True)
        nuove, problemi = importa(cfg, stato, args.reimporta)
        SPESE_JSON.write_text(json.dumps(stato, ensure_ascii=False, indent=2), encoding="utf-8")
        print(f"{nuove} nuove, {len(stato['fatture'])} totali in registro")

    # 3) report
    print("[3/3] Generazione Excel e dashboard...", end=" ", flush=True)
    an = analizza(cfg, stato)
    scrivi_excel(cfg, stato, an)
    scrivi_html(cfg, stato, an)
    print("fatto\n")

    print("--- Situazione ---")
    print(f"Costo ricorrente:   {an['ricorrente_mensile']:.2f} EUR/mese "
          f"({an['ricorrente_annuale']:.2f} EUR/anno)")
    print(f"Previsione 12 mesi: {an['previsione_totale_12m']:.2f} EUR")
    prossima = next((s for s in an["scadenze"] if s["giorni"] >= 0), None)
    if prossima:
        segno = "!!" if prossima["giorni"] <= 45 else "  "
        print(f"{segno} Prossima scadenza: {prossima['data']} — {prossima['contratto']} "
              f"(tra {prossima['giorni']} giorni)")
        if prossima.get("aumento_eur"):
            print(f"   Impatto: +{prossima['aumento_eur']:.2f} EUR")
    be = an.get("breakeven")
    if be and be.get("costo_per_organizzazione") is not None:
        print(f"Costo per organizzazione: {be['costo_per_organizzazione']:.2f} EUR/mese "
              f"({be['organizzazioni_totali']} org, {be['utenti_attivi']} utenti attivi)")

    if problemi:
        print("\n--- Da controllare ---")
        for p in problemi:
            print(f"  - {p}")

    print(f"\nGenerati:\n  {XLSX_OUT}\n  {HTML_OUT}")

    if args.dashboard:
        print()
        sys.path.insert(0, str(SISTEMA))
        import server
        return server.avvia(args.porta, apri_browser=True)

    return 0


if __name__ == "__main__":
    sys.exit(main())
